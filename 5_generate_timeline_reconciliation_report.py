import argparse
import sys
import time
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from timeline_reconciliation_common import OUTPUT_DIR, connect_motherduck, parse_config_txt


DEFAULT_OUTPUT_BASENAME = "timeline_reconciliation_report"


def default_report_output_path() -> Path:
    ts = time.strftime("%Y%m%d_%H%M%S")
    return OUTPUT_DIR / f"{DEFAULT_OUTPUT_BASENAME}_{ts}.xlsx"

HEADERS_LINKS = [
    "#",
    "TimelineName",
    "TaskRowId",
    "TaskCode",
    "TaskName",
    "WbsName",
    "TaskClass",
    "TaskClassConfidence",
    "TaskClassReason",
    "LinkRank",
    "LinkReason",
    "MdrDocumentTitle",
    "DocumentRaciTitle",
    "Resolver LLM — Candidato Top 1",
    "Resolver LLM — Candidato Top 2",
    "Resolver LLM — Candidato Top 3",
    "Resolver LLM — Candidato Top 4",
    "Resolver LLM — Candidato Top 5",
]

DATE_HEADERS_SELECTED = [
    "Data inizio (actualized)",
    "Data fine (actualized)",
]
DATE_ROW_KEYS_SELECTED = [
    "SelectedStartDate",
    "SelectedFinishDate",
]
# Grezze usate nel COALESCE actualized (Actual → Early → Target)
DATE_HEADERS_RAW = [
    "EarlyStartDate",
    "EarlyEndDate",
    "ActualStartDate",
    "ActualEndDate",
    "TargetStartDate",
    "TargetEndDate",
]
HEADERS_LINKS = HEADERS_LINKS + DATE_HEADERS_SELECTED + DATE_HEADERS_RAW

COL_WIDTHS_LINKS = [
    6,
    28,
    10,
    16,
    44,
    30,
    12,
    14,
    42,
    9,
    52,
    42,
    38,
    40,
    40,
    40,
    40,
    40,
]
COL_WIDTHS_LINKS = COL_WIDTHS_LINKS + [22, 22] + [14] * 6
HEADERS_TASKS = [
    "#",
    "TimelineName",
    "TaskRowId",
    "TaskCode",
    "TaskName",
    "WbsName",
    "TaskClass",
    "TaskClassConfidence",
    "TaskClassReason",
]
HEADERS_TASKS = HEADERS_TASKS + DATE_HEADERS_SELECTED + DATE_HEADERS_RAW + ["ResolverLinkCount"]
COL_WIDTHS_TASKS = [6, 30, 10, 16, 54, 36, 12, 14, 52] + [22, 22] + [14] * 6 + [12]
HEADERS_CLASSIFY = [
    "#",
    "TimelineName",
    "TaskRowId",
    "TaskCode",
    "TaskName",
    "WbsName",
    "TaskClass",
    "TaskClassConfidence",
    "TaskClassReason",
]
COL_WIDTHS_CLASSIFY = [6, 30, 10, 16, 58, 40, 12, 14, 60]

NAVY = "0D1B2A"
WHITE = "FFFFFF"
GRID = "CBD5E1"

CLASS_BG = {"ENG_DOC": "DBEAFE", "OTHER": "F3F4F6"}
CLASS_FG = {"ENG_DOC": "1E3A8A", "OTHER": "374151"}
SECTION_COLORS = {
    "task": ("1E40AF", "EFF6FF"),
    "resolver": ("166534", "ECFDF5"),
    "raci": ("6D28D9", "F5F3FF"),
    "llm_top": ("C2410C", "FFF7ED"),
    "selected": ("0F766E", "CCFBF1"),
    "dates": ("64748B", "F1F5F9"),
}

thin = Side(style="thin", color=GRID)


def _fill(color):
    return PatternFill("solid", start_color=color)


def _align(center=False):
    return Alignment(horizontal=("center" if center else "left"), vertical="top", wrap_text=True)


def _border():
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _fmt_ts(value):
    if value is None:
        return ""
    text = str(value)
    return text.replace("T", " ")


def _safe_text(value):
    if value is None:
        return ""
    return str(value)[:32767]


def _is_na_scalar(val):
    if val is None:
        return True
    try:
        return bool(pd.isna(val))
    except (TypeError, ValueError):
        return False


def _safe_int(val, default=None):
    """int(...) che non esplode su NA / NaN / NAType."""
    if _is_na_scalar(val):
        return default
    try:
        return int(val)
    except (TypeError, ValueError):
        return default


def _link_rank_cell(val):
    """LinkRank per Excel: stringa vuota se assente (righe senza link)."""
    i = _safe_int(val, default=None)
    return "" if i is None else i


def _fmt_resolver_llm_slot(title, why):
    """Cell text for one LLM shortlist candidate (title + why_plausible only)."""
    title = (title or "").strip() or "—"
    why_str = (why or "").strip()
    if not why_str:
        return f"TITOLO (RACI / MDR):\n{title}\n\nNOTA:\n—"
    return f"TITOLO (RACI / MDR):\n{title}\n\nNOTA:\n{why_str}"


def load_llm_top_lookup(conn, db_name, embedding_model, timeline_name=None):
    """
    (TimelineName, TaskRowId) -> list of 5 formatted strings (or None per empty slot).
    """
    timeline_filter = ""
    params = [embedding_model]
    if timeline_name:
        timeline_filter = "AND TimelineName = ?"
        params.append(timeline_name)
    sql = f"""
    SELECT
        TimelineName,
        TaskRowId,
        CandidateRankWithinResolver,
        COALESCE(NULLIF(TRIM(ConsolidatedRaciTitle), ''), NULLIF(TRIM(MdrDocumentTitle), '')) AS display_title,
        WhyPlausible
    FROM {db_name}.timeline_reconciliation.TimelineTaskToMdrResolverLlmTopCandidates
    WHERE EmbeddingModel = ?
      AND CandidateRankWithinResolver >= 1
      AND CandidateRankWithinResolver <= 5
    {timeline_filter}
    ORDER BY TimelineName, TaskRowId, CandidateRankWithinResolver
    """
    try:
        df = conn.execute(sql, params).fetchdf()
    except Exception:
        return {}
    lookup = {}
    for _, row in df.iterrows():
        tid = _safe_int(row.get("TaskRowId"), default=None)
        if tid is None:
            continue
        key = (_safe_text(row.get("TimelineName")), tid)
        rank = _safe_int(row.get("CandidateRankWithinResolver"), default=None)
        if rank is None or rank < 1 or rank > 5:
            continue
        txt = _fmt_resolver_llm_slot(row.get("display_title"), row.get("WhyPlausible"))
        if key not in lookup:
            lookup[key] = [None] * 5
        lookup[key][rank - 1] = txt
    return lookup


def load_links_rows(conn, db_name, embedding_model, timeline_name=None, llm_top_lookup=None):
    timeline_filter = ""
    params = []
    if timeline_name:
        timeline_filter = "WHERE c.TimelineName = ?"
        params.append(timeline_name)

    sql = f"""
    WITH classified_latest AS (
        SELECT *
        FROM (
            SELECT
                c.*,
                ROW_NUMBER() OVER (
                    PARTITION BY c.TimelineName, c.TaskRowId
                    ORDER BY c.UpdatedAt DESC, c.CreatedAt DESC
                ) AS rn
            FROM {db_name}.timeline_reconciliation.TimelineTasksClassified c
        ) x
        WHERE x.rn = 1
    ),
    links_latest AS (
        SELECT *
        FROM (
            SELECT
                l.*,
                ROW_NUMBER() OVER (
                    PARTITION BY l.TimelineName, l.TaskRowId, l.MdrTitleKey
                    ORDER BY l.CreatedAt DESC
                ) AS rn
            FROM {db_name}.timeline_reconciliation.TimelineTaskToMdrLinks l
        ) y
        WHERE y.rn = 1
    ),
    link_counts AS (
        SELECT
            TimelineName,
            TaskRowId,
            COUNT(*) AS ResolverLinkCount
        FROM links_latest
        GROUP BY TimelineName, TaskRowId
    )
    SELECT
        c.TimelineName,
        c.ProjectCode,
        c.TaskRowId,
        c.TaskCode,
        c.TaskName,
        c.WbsName,
        c.TaskClass,
        c.TaskClassConfidence,
        c.TaskClassReason,
        COALESCE(lc.ResolverLinkCount, 0) AS ResolverLinkCount,
        l.MdrDocumentTitle AS LinkMdrDocumentTitle,
        l.ConsolidatedRaciTitle AS DocumentRaciTitle,
        l.LinkReason,
        l.LinkRank,
        COALESCE(
            l.SelectedStartDate,
            COALESCE(c.ActualStartDate, c.EarlyStartDate, c.TargetStartDate)
        ) AS SelectedStartDate,
        COALESCE(
            l.SelectedFinishDate,
            COALESCE(c.ActualEndDate, c.EarlyEndDate, c.TargetEndDate)
        ) AS SelectedFinishDate,
        COALESCE(l.EarlyStartDate, c.EarlyStartDate) AS EarlyStartDate,
        COALESCE(l.EarlyEndDate, c.EarlyEndDate) AS EarlyEndDate,
        COALESCE(l.ActualStartDate, c.ActualStartDate) AS ActualStartDate,
        COALESCE(l.ActualEndDate, c.ActualEndDate) AS ActualEndDate,
        COALESCE(l.TargetStartDate, c.TargetStartDate) AS TargetStartDate,
        COALESCE(l.TargetEndDate, c.TargetEndDate) AS TargetEndDate
    FROM classified_latest c
    LEFT JOIN link_counts lc
      ON lc.TimelineName = c.TimelineName
     AND lc.TaskRowId = c.TaskRowId
    LEFT JOIN links_latest l
      ON l.TimelineName = c.TimelineName
     AND l.TaskRowId = c.TaskRowId
    {timeline_filter}
    ORDER BY c.TimelineName, c.TaskRowId, l.LinkRank
    """
    df = conn.execute(sql, params).fetchdf()

    rows = []
    for _, row in df.iterrows():
        tid = _safe_int(row.get("TaskRowId"), default=None)
        if tid is None:
            continue
        out = {
            "TimelineName": _safe_text(row.get("TimelineName")),
            "ProjectCode": _safe_text(row.get("ProjectCode")),
            "TaskRowId": tid,
            "TaskCode": _safe_text(row.get("TaskCode")),
            "TaskName": _safe_text(row.get("TaskName")),
            "WbsName": _safe_text(row.get("WbsName")),
            "TaskClass": _safe_text(row.get("TaskClass")),
            "TaskClassConfidence": _safe_text(row.get("TaskClassConfidence")),
            "TaskClassReason": _safe_text(row.get("TaskClassReason")),
            "ResolverLinkCount": _safe_int(row.get("ResolverLinkCount"), default=0) or 0,
            "LinkRank": _link_rank_cell(row.get("LinkRank")),
            "LinkReason": _safe_text(row.get("LinkReason")),
            "MdrDocumentTitle": _safe_text(row.get("LinkMdrDocumentTitle")),
            "DocumentRaciTitle": _safe_text(row.get("DocumentRaciTitle")),
        }
        for header, key in zip(DATE_HEADERS_SELECTED, DATE_ROW_KEYS_SELECTED):
            out[header] = _fmt_ts(row.get(key))
        for col in DATE_HEADERS_RAW:
            out[col] = _fmt_ts(row.get(col))
        if llm_top_lookup is not None:
            key = (out["TimelineName"], out["TaskRowId"])
            slots = llm_top_lookup.get(key, [None] * 5)
            for i in range(5):
                out[f"ResolverLlmTop{i + 1}"] = slots[i] if i < len(slots) else None
        else:
            for i in range(5):
                out[f"ResolverLlmTop{i + 1}"] = None
        rows.append(out)
    return rows


def build_task_summary_rows(link_rows):
    grouped = {}
    for r in link_rows:
        key = (r["TimelineName"], r["TaskRowId"])
        if key not in grouped:
            grouped[key] = {
                "TimelineName": r["TimelineName"],
                "TaskRowId": r["TaskRowId"],
                "TaskCode": r["TaskCode"],
                "TaskName": r["TaskName"],
                "WbsName": r["WbsName"],
                "TaskClass": r["TaskClass"],
                "TaskClassConfidence": r["TaskClassConfidence"],
                "TaskClassReason": r["TaskClassReason"],
                "ResolverLinkCount": r["ResolverLinkCount"],
            }
            for header in DATE_HEADERS_SELECTED:
                grouped[key][header] = r.get(header, "")
            for col in DATE_HEADERS_RAW:
                grouped[key][col] = r.get(col, "")
    return sorted(grouped.values(), key=lambda x: (x["TimelineName"], x["TaskRowId"]))


def _links_selected_col_range():
    start = HEADERS_LINKS.index(DATE_HEADERS_SELECTED[0]) + 1
    end = HEADERS_LINKS.index(DATE_HEADERS_SELECTED[-1]) + 1
    return start, end


def _links_raw_col_range():
    start = HEADERS_LINKS.index(DATE_HEADERS_RAW[0]) + 1
    end = HEADERS_LINKS.index(DATE_HEADERS_RAW[-1]) + 1
    return start, end


def _links_col_section(col_idx):
    if col_idx <= 9:
        return "task"
    if col_idx <= 13:
        return "resolver"
    if col_idx <= 18:
        return "llm_top"
    sel_start, sel_end = _links_selected_col_range()
    if sel_start <= col_idx <= sel_end:
        return "selected"
    return "dates"


def _is_selected_header(header):
    return header in DATE_HEADERS_SELECTED


def _build_links_sheet(ws, title, rows):
    ws.merge_cells(f"A1:{get_column_letter(len(HEADERS_LINKS))}1")
    ws["A1"] = f"Timeline Reconciliation Report - {title} | rows: {len(rows)}"
    ws["A1"].font = Font(name="Arial", bold=True, size=11, color=WHITE)
    ws["A1"].fill = _fill(NAVY)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    ws.row_dimensions[1].height = 24

    sel_start, sel_end = _links_selected_col_range()
    raw_start, raw_end = _links_raw_col_range()
    sections = [
        ("Task + Classify", 2, 9, "task"),
        ("Resolver Final Link", 10, 13, "resolver"),
        ("Resolver LLM shortlist (max 5)", 14, 18, "llm_top"),
        ("Date actualized (Actual → Early → Target)", sel_start, sel_end, "selected"),
        ("Grezze input actualized", raw_start, raw_end, "dates"),
    ]
    ws.cell(row=2, column=1, value="#")
    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
    cell = ws.cell(row=2, column=1)
    cell.font = Font(name="Arial", bold=True, size=9, color=WHITE)
    cell.fill = _fill("1A2E42")
    cell.alignment = _align(center=True)
    cell.border = _border()

    for label, start, end, key in sections:
        header_color, _ = SECTION_COLORS[key]
        ws.merge_cells(start_row=2, start_column=start, end_row=2, end_column=end)
        sc = ws.cell(row=2, column=start, value=label)
        sc.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        sc.fill = _fill(header_color)
        sc.alignment = _align(center=True)
        for col in range(start, end + 1):
            ws.cell(row=2, column=col).border = _border()

    for idx, (header, width) in enumerate(zip(HEADERS_LINKS, COL_WIDTHS_LINKS), 1):
        if idx == 1:
            ws.column_dimensions[get_column_letter(idx)].width = width
            continue
        c = ws.cell(row=3, column=idx, value=header)
        section_key = _links_col_section(idx)
        header_color, _ = SECTION_COLORS[section_key]
        header_size = 10 if _is_selected_header(header) else 9
        c.font = Font(name="Arial", bold=True, size=header_size, color=WHITE)
        c.fill = _fill(header_color)
        c.alignment = _align(center=True)
        c.border = _border()
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[3].height = 32

    for i, row in enumerate(rows, 1):
        excel_row = i + 3
        task_class = row.get("TaskClass", "")
        class_bg = CLASS_BG.get(task_class, "F8FAFC")
        fg = CLASS_FG.get(task_class, "1A1A2E")

        values = [
            i,
            row["TimelineName"],
            row["TaskRowId"],
            row["TaskCode"],
            row["TaskName"],
            row["WbsName"],
            row["TaskClass"],
            row["TaskClassConfidence"],
            row["TaskClassReason"],
            row["LinkRank"],
            row["LinkReason"],
            row["MdrDocumentTitle"],
            row["DocumentRaciTitle"],
            row.get("ResolverLlmTop1") or "—",
            row.get("ResolverLlmTop2") or "—",
            row.get("ResolverLlmTop3") or "—",
            row.get("ResolverLlmTop4") or "—",
            row.get("ResolverLlmTop5") or "—",
        ]
        for col in DATE_HEADERS_SELECTED + DATE_HEADERS_RAW:
            values.append(row.get(col, ""))
        date_col_start = len(HEADERS_LINKS) - len(DATE_HEADERS_SELECTED) - len(DATE_HEADERS_RAW) + 1
        sel_start, sel_end = _links_selected_col_range()
        for col_idx, value in enumerate(values, 1):
            c = ws.cell(row=excel_row, column=col_idx, value=_safe_text(value))
            c.border = _border()
            c.alignment = _align(
                center=col_idx in (1, 3, 7, 8, 10) or col_idx >= date_col_start
            )
            if col_idx in (7, 8):
                c.fill = _fill(class_bg)
                c.font = Font(name="Arial", bold=True, size=9, color=fg)
            elif sel_start <= col_idx <= sel_end:
                _, cell_color = SECTION_COLORS["selected"]
                c.fill = _fill(cell_color)
                c.font = Font(name="Arial", bold=True, size=11, color="064E3B")
            else:
                _, cell_color = SECTION_COLORS[_links_col_section(col_idx)]
                c.fill = _fill(cell_color)
                if 14 <= col_idx <= 18:
                    c.font = Font(name="Arial", size=9, color="7C2D12")
                else:
                    c.font = Font(name="Arial", size=9, color="1A1A2E")

        ws.row_dimensions[excel_row].height = 72

    ws.auto_filter.ref = f"A3:{get_column_letter(len(HEADERS_LINKS))}{len(rows) + 3}"
    ws.freeze_panes = "A4"


def _build_tasks_sheet(ws, title, rows):
    ws.merge_cells(f"A1:{get_column_letter(len(HEADERS_TASKS))}1")
    ws["A1"] = f"Timeline Task Summary - {title} | tasks: {len(rows)}"
    ws["A1"].font = Font(name="Arial", bold=True, size=11, color=WHITE)
    ws["A1"].fill = _fill(NAVY)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    ws.row_dimensions[1].height = 24

    sel_start = HEADERS_TASKS.index(DATE_HEADERS_SELECTED[0]) + 1
    sel_end = HEADERS_TASKS.index(DATE_HEADERS_SELECTED[-1]) + 1
    sel_header_color, _ = SECTION_COLORS["selected"]

    for idx, (header, width) in enumerate(zip(HEADERS_TASKS, COL_WIDTHS_TASKS), 1):
        c = ws.cell(row=2, column=idx, value=header)
        if sel_start <= idx <= sel_end:
            c.fill = _fill(sel_header_color)
            c.font = Font(name="Arial", bold=True, size=10, color=WHITE)
        else:
            c.fill = _fill("1A2E42")
            c.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        c.alignment = _align(center=True)
        c.border = _border()
        ws.column_dimensions[get_column_letter(idx)].width = width

    for i, row in enumerate(rows, 1):
        excel_row = i + 2
        task_class = row.get("TaskClass", "")
        bg = CLASS_BG.get(task_class, "F8FAFC")
        fg = CLASS_FG.get(task_class, "1A1A2E")
        values = [i] + [row[h] for h in HEADERS_TASKS[1:]]
        date_start = HEADERS_TASKS.index(DATE_HEADERS_SELECTED[0]) + 1
        _, sel_cell_color = SECTION_COLORS["selected"]
        for col_idx, value in enumerate(values, 1):
            c = ws.cell(row=excel_row, column=col_idx, value=_safe_text(value))
            c.border = _border()
            c.alignment = _align(
                center=col_idx in (1, 3, 7, 8, len(HEADERS_TASKS)) or col_idx >= date_start
            )
            if col_idx in (7, 8):
                c.fill = _fill(bg)
                c.font = Font(name="Arial", bold=True, size=9, color=fg)
            elif sel_start <= col_idx <= sel_end:
                c.fill = _fill(sel_cell_color)
                c.font = Font(name="Arial", bold=True, size=11, color="064E3B")
            else:
                c.fill = _fill("FFFFFF")
                c.font = Font(name="Arial", size=9, color="1A1A2E")
        ws.row_dimensions[excel_row].height = 60

    ws.auto_filter.ref = f"A2:{get_column_letter(len(HEADERS_TASKS))}{len(rows) + 2}"
    ws.freeze_panes = "A3"


def _build_classify_sheet(ws, title, rows):
    ws.merge_cells(f"A1:{get_column_letter(len(HEADERS_CLASSIFY))}1")
    ws["A1"] = f"Timeline Task Classification - {title} | tasks: {len(rows)}"
    ws["A1"].font = Font(name="Arial", bold=True, size=11, color=WHITE)
    ws["A1"].fill = _fill(NAVY)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    ws.row_dimensions[1].height = 24

    for idx, (header, width) in enumerate(zip(HEADERS_CLASSIFY, COL_WIDTHS_CLASSIFY), 1):
        c = ws.cell(row=2, column=idx, value=header)
        c.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        c.fill = _fill("1A2E42")
        c.alignment = _align(center=True)
        c.border = _border()
        ws.column_dimensions[get_column_letter(idx)].width = width

    for i, row in enumerate(rows, 1):
        excel_row = i + 2
        task_class = row.get("TaskClass", "")
        bg = CLASS_BG.get(task_class, "F8FAFC")
        fg = CLASS_FG.get(task_class, "1A1A2E")
        values = [
            i,
            row["TimelineName"],
            row["TaskRowId"],
            row["TaskCode"],
            row["TaskName"],
            row["WbsName"],
            row["TaskClass"],
            row["TaskClassConfidence"],
            row["TaskClassReason"],
        ]
        for col_idx, value in enumerate(values, 1):
            c = ws.cell(row=excel_row, column=col_idx, value=_safe_text(value))
            c.border = _border()
            c.alignment = _align(center=col_idx in (1, 3, 7, 8))
            if col_idx in (7, 8):
                c.fill = _fill(bg)
                c.font = Font(name="Arial", bold=True, size=9, color=fg)
            else:
                c.fill = _fill("FFFFFF")
                c.font = Font(name="Arial", size=9, color="1A1A2E")
        ws.row_dimensions[excel_row].height = 58

    ws.auto_filter.ref = f"A2:{get_column_letter(len(HEADERS_CLASSIFY))}{len(rows) + 2}"
    ws.freeze_panes = "A3"


def _save_workbook(wb, output_path: Path) -> Path:
    output_path = Path(output_path).resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(output_path)
        return output_path
    except PermissionError:
        fallback = default_report_output_path()
        wb.save(fallback)
        print(
            f"ATTENZIONE: impossibile scrivere {output_path} (file aperto o permessi insufficienti).\n"
            f"Report salvato in: {fallback}",
            file=sys.stderr,
        )
        return fallback


def write_report(link_rows, output_path):
    task_rows = build_task_summary_rows(link_rows)
    wb = Workbook()
    ws_classify = wb.active
    ws_classify.title = "Task + Classify"
    _build_classify_sheet(ws_classify, "Task + Classify", task_rows)

    eng_link_rows = [r for r in link_rows if r.get("TaskClass") == "ENG_DOC"]
    _build_links_sheet(wb.create_sheet("ENG_DOC Full"), "ENG_DOC Full", eng_link_rows)

    return _save_workbook(wb, output_path)


def main():
    parser = argparse.ArgumentParser(description="Generate timeline reconciliation Excel report (classify + resolver).")
    parser.add_argument("--timeline", default="", help="Optional TimelineName filter.")
    parser.add_argument(
        "--output",
        default="",
        help="Output xlsx path (default: output/timeline_reconciliation_report_YYYYMMDD_HHMMSS.xlsx).",
    )
    parser.add_argument(
        "--embedding-model",
        default="",
        help="EmbeddingModel per TimelineTaskToMdrResolverLlmTopCandidates (default: config EMBEDDING_MODEL).",
    )
    args = parser.parse_args()

    cfg = parse_config_txt()
    db_name = cfg.get("MOTHERDUCK_DB", "my_db").strip() or "my_db"
    embedding_model = (args.embedding_model.strip() or cfg.get("EMBEDDING_MODEL") or "text-embedding-3-small").strip()
    timeline_name = args.timeline.strip() or None
    output_path = Path(args.output).resolve() if args.output.strip() else default_report_output_path()

    llm_lookup = {}
    try:
        conn = connect_motherduck(cfg)
    except Exception as exc:
        err = str(exc).lower()
        hint = ""
        if "motherduck" in err or "connection" in err or "download" in err or "establish" in err:
            hint = (
                " Suggerimento: controlla la connessione internet, firewall/proxy e che MotherDuck sia "
                "raggiungibile; riprova tra qualche minuto."
            )
        print(
            f"ERRORE: connessione al database MotherDuck fallita.{hint}\n",
            f"Dettaglio tecnico: {exc}",
            file=sys.stderr,
        )
        raise SystemExit(1) from exc
    try:
        llm_lookup = load_llm_top_lookup(conn, db_name, embedding_model, timeline_name=timeline_name)
        rows = load_links_rows(
            conn,
            db_name,
            embedding_model,
            timeline_name=timeline_name,
            llm_top_lookup=llm_lookup,
        )
    finally:
        conn.close()

    saved_path = write_report(rows, output_path)
    print(f"[OK] Report generated: {saved_path}")
    task_rows = build_task_summary_rows(rows)
    print(f"Total link-view rows: {len(rows)}")
    print(f"Total tasks: {len(task_rows)}")
    print(f"ENG_DOC tasks: {sum(1 for r in task_rows if r.get('TaskClass') == 'ENG_DOC')}")
    print(f"OTHER tasks: {sum(1 for r in task_rows if r.get('TaskClass') == 'OTHER')}")
    print(f"Tasks with links: {sum(1 for r in task_rows if r.get('ResolverLinkCount', 0) > 0)}")
    print(f"Tasks with resolver LLM shortlist in DB: {len(llm_lookup)} (EmbeddingModel={embedding_model})")


if __name__ == "__main__":
    main()
